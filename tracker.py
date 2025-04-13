from fastapi import FastAPI, Depends, HTTPException,Request,Form
from sqlalchemy import create_engine, Column, Integer, String, BigInteger, ForeignKey, DateTime,Index
from sqlalchemy.orm import sessionmaker, declarative_base, relationship, Session
from sqlalchemy.ext.declarative import declarative_base
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr
from apscheduler.schedulers.background import BackgroundScheduler
import datetime
import bcrypt
import requests
from urllib.parse import quote_plus,urlparse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
import xlsxwriter
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime,timedelta
from fastapi import APIRouter
from fastapi import Form
from fastapi import Body
from starlette.staticfiles import StaticFiles
from starlette.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
import logging
from iso3166 import countries
app = FastAPI()
import pycountry
if not os.path.exists("static"):
    os.makedirs("static")

templates = Jinja2Templates(directory="templates")
# Database Configuration
DB_USER = "user_name"
DB_PASSWORD = " password"
DB_HOST = "localhost"
MASTER_DB = "db_name"

# Encode password to handle special characters
encoded_password = quote_plus(DB_PASSWORD)
DATABASE_URL = f"mysql+mysqlconnector://{DB_USER}:{encoded_password}@{DB_HOST}/{MASTER_DB}"

engine = create_engine(DATABASE_URL, echo=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()
# Password Hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ORM Models
class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(255), nullable=False)
    email_id = Column(String(255), unique=True, index=True, nullable=False)
    phone_number = Column(BigInteger, unique=True, nullable=True)
    password_hash = Column(String(255), nullable=False)
    login_id = Column(String(255), unique=True, index=True, nullable=False)
    projects = relationship("Project", back_populates="user")

class Project(Base):
    __tablename__ = "projects"
    id = Column(Integer, primary_key=True, index=True)
    project_name = Column(String(255), index=True, nullable=False)
    project_description = Column(String(255), nullable=True)
    user_id = Column(Integer, ForeignKey("users.id"),nullable=False)
    country = Column(String(10), nullable=True)
    user = relationship("User", back_populates="projects")
    urls = relationship("URL", back_populates="project")

class URL(Base):
    __tablename__ = "urls"
    id = Column(Integer, primary_key=True, index=True)
    url = Column(String, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"))
    country = Column(String, nullable=False)
    project = relationship("Project", back_populates="urls")
    ranks = relationship("Rank", back_populates="url")
    keywords = relationship("Keyword", back_populates="url")
    __table_args__ = (Index('idx_unique_url_per_project', 'url', 'project_id', unique=True),)

class Keyword(Base):
    __tablename__ = "keywords"
    id = Column(Integer, primary_key=True, index=True)
    keyword = Column(String, index=True)
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"))
    
    url_id = Column(Integer, ForeignKey("urls.id", ondelete="CASCADE"), nullable=False)
    
    url = relationship("URL", back_populates="keywords") 
    ranks = relationship("Rank", back_populates="keyword")

    __table_args__ = (Index('idx_unique_keyword_per_project', 'keyword', 'project_id', unique=True),)

class Rank(Base):
    __tablename__ = "ranks"
    id = Column(Integer, primary_key=True, index=True)
    url_id = Column(Integer, ForeignKey("urls.id", ondelete="CASCADE"))
    keyword_id = Column(Integer, ForeignKey("keywords.id", ondelete="CASCADE"))
    project_id = Column(Integer, ForeignKey("projects.id", ondelete="CASCADE"))
    ranks = Column(Integer)
    page_number = Column(Integer)
    country = Column(String, nullable=False)
    date = Column(DateTime, default=datetime.utcnow)

    url = relationship("URL", back_populates="ranks")
    keyword = relationship("Keyword", back_populates="ranks")

    __table_args__ = (Index('idx_url_keyword_project', 'url_id', 'keyword_id', 'project_id', unique=True),)
# Pydantic Schemas
class UserCreate(BaseModel):
    name: str
    email_id: EmailStr
    phone_number: int
    password: str

class LoginSchema(BaseModel):
    email_id: EmailStr
    password: str

class ProjectCreate(BaseModel):
    project_name: str
    project_description: str
    user_id:int


class TrackRankRequest(BaseModel):
    project_id: int
    urls: list[str]
    keywords: list[str]
    country: str

# FastAPI App
app = FastAPI()
bcrypt._about_ = {"_version_": "4.0.1"}

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()
OUTPUT_DIR = "output_files"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Single SERPAPI Key
SERP_API_KEY = "Serp_api_key"

# Country Mapping
country_domains = {
    "afghanistan": "google.com.af",
    "albania": "google.al",
    "algeria": "google.dz",
    "andorra": "google.ad",
    "angola": "google.co.ao",
    "argentina": "google.com.ar",
    "armenia": "google.am",
    "australia": "google.com.au",
    "austria": "google.at",
    "azerbaijan": "google.az",
    "bahrain": "google.com.bh",
    "bangladesh": "google.com.bd",
    "belarus": "google.by",
    "belgium": "google.be",
    "belize": "google.com.bz",
    "benin": "google.bj",
    "bhutan": "google.bt",
    "bolivia": "google.com.bo",
    "bosnia and herzegovina": "google.ba",
    "botswana": "google.co.bw",
    "brazil": "google.com.br",
    "brunei": "google.com.bn",
    "bulgaria": "google.bg",
    "burkina faso": "google.bf",
    "burundi": "google.bi",
    "cambodia": "google.com.kh",
    "cameroon": "google.cm",
    "canada": "google.ca",
    "cape verde": "google.cv",
    "chad": "google.td",
    "chile": "google.cl",
    "china": "google.cn",
    "colombia": "google.com.co",
    "congo": "google.cg",
    "costa rica": "google.co.cr",
    "croatia": "google.hr",
    "cuba": "google.com.cu",
    "cyprus": "google.com.cy",
    "czech republic": "google.cz",
    "denmark": "google.dk",
    "djibouti": "google.dj",
    "dominican republic": "google.com.do",
    "ecuador": "google.com.ec",
    "egypt": "google.com.eg",
    "el salvador": "google.com.sv",
    "estonia": "google.ee",
    "ethiopia": "google.com.et",
    "fiji": "google.com.fj",
    "finland": "google.fi",
    "france": "google.fr",
    "gabon": "google.ga",
    "gambia": "google.gm",
    "georgia": "google.ge",
    "germany": "google.de",
    "ghana": "google.com.gh",
    "greece": "google.gr",
    "guatemala": "google.com.gt",
    "guinea": "google.gn",
    "haiti": "google.ht",
    "honduras": "google.hn",
    "hong kong": "google.com.hk",
    "hungary": "google.hu",
    "iceland": "google.is",
    "india": "google.co.in",
    "indonesia": "google.co.id",
    "iran": "google.com.ir",
    "iraq": "google.iq",
    "ireland": "google.ie",
    "israel": "google.co.il",
    "italy": "google.it",
    "jamaica": "google.com.jm",
    "japan": "google.co.jp",
    "jordan": "google.jo",
    "kazakhstan": "google.kz",
    "kenya": "google.co.ke",
    "kuwait": "google.com.kw",
    "kyrgyzstan": "google.kg",
    "laos": "google.la",
    "latvia": "google.lv",
    "lebanon": "google.com.lb",
    "libya": "google.com.ly",
    "liechtenstein": "google.li",
    "lithuania": "google.lt",
    "luxembourg": "google.lu",
    "macedonia": "google.mk",
    "madagascar": "google.mg",
    "malawi": "google.mw",
    "malaysia": "google.com.my",
    "maldives": "google.mv",
    "mali": "google.ml",
    "malta": "google.com.mt",
    "mauritania": "google.mr",
    "mauritius": "google.mu",
    "mexico": "google.com.mx",
    "moldova": "google.md",
    "monaco": "google.mc",
    "mongolia": "google.mn",
    "montenegro": "google.me",
    "morocco": "google.co.ma",
    "mozambique": "google.co.mz",
    "myanmar": "google.com.mm",
    "namibia": "google.com.na",
    "nepal": "google.com.np",
    "netherlands": "google.nl",
    "new zealand": "google.co.nz",
    "nicaragua": "google.com.ni",
    "niger": "google.ne",
    "nigeria": "google.com.ng",
    "norway": "google.no",
    "oman": "google.com.om",
    "pakistan": "google.com.pk",
    "panama": "google.com.pa",
    "paraguay": "google.com.py",
    "peru": "google.com.pe",
    "philippines": "google.com.ph",
    "poland": "google.pl",
    "portugal": "google.pt",
    "puerto rico": "google.com.pr",
    "qatar": "google.com.qa",
    "romania": "google.ro",
    "russia": "google.ru",
    "rwanda": "google.rw",
    "saudi arabia": "google.com.sa",
    "senegal": "google.sn",
    "serbia": "google.rs",
    "singapore": "google.com.sg",
    "slovakia": "google.sk",
    "slovenia": "google.si",
    "south africa": "google.co.za",
    "south korea": "google.co.kr",
    "spain": "google.es",
    "sri lanka": "google.lk",
    "sudan": "google.com.sd",
    "sweden": "google.se",
    "switzerland": "google.ch",
    "syria": "google.sy",
    "taiwan": "google.com.tw",
    "tajikistan": "google.com.tj",
    "tanzania": "google.co.tz",
    "thailand": "google.co.th",
    "tunisia": "google.tn",
    "turkey": "google.com.tr",
    "turkmenistan": "google.tm",
    "uganda": "google.co.ug",
    "ukraine": "google.com.ua",
    "united arab emirates": "google.ae",
    "united kingdom": "google.co.uk",
    "united states": "google.com",
    "uruguay": "google.com.uy",
    "uzbekistan": "google.co.uz",
    "venezuela": "google.co.ve",
    "vietnam": "google.com.vn",
    "yemen": "google.com.ye",
    "zambia": "google.co.zm",
    "zimbabwe": "google.co.zw"
}

country_codes = {
    "Afghanistan": "af",
    "Albania": "al",
    "Algeria": "dz",
    "Andorra": "ad",
    "Angola": "ao",
    "Argentina": "ar",
    "Armenia": "am",
    "Australia": "au",
    "Austria": "at",
    "Azerbaijan": "az",
    "Bahamas": "bs",
    "Bahrain": "bh",
    "Bangladesh": "bd",
    "Barbados": "bb",
    "Belarus": "by",
    "Belgium": "be",
    "Belize": "bz",
    "Benin": "bj",
    "Bhutan": "bt",
    "Bolivia": "bo",
    "Bosnia and Herzegovina": "ba",
    "Botswana": "bw",
    "Brazil": "br",
    "Brunei": "bn",
    "Bulgaria": "bg",
    "Burkina Faso": "bf",
    "Burundi": "bi",
    "Cambodia": "kh",
    "Cameroon": "cm",
    "Canada": "ca",
    "Cape Verde": "cv",
    "Central African Republic": "cf",
    "Chad": "td",
    "Chile": "cl",
    "China": "cn",
    "Colombia": "co",
    "Comoros": "km",
    "Congo": "cg",
    "Costa Rica": "cr",
    "Croatia": "hr",
    "Cuba": "cu",
    "Cyprus": "cy",
    "Czechia": "cz",
    "Denmark": "dk",
    "Djibouti": "dj",
    "Dominica": "dm",
    "Dominican Republic": "do",
    "Ecuador": "ec",
    "Egypt": "eg",
    "El Salvador": "sv",
    "Equatorial Guinea": "gq",
    "Eritrea": "er",
    "Estonia": "ee",
    "Eswatini": "sz",
    "Ethiopia": "et",
    "Fiji": "fj",
    "Finland": "fi",
    "France": "fr",
    "Gabon": "ga",
    "Gambia": "gm",
    "Georgia": "ge",
    "Germany": "de",
    "Ghana": "gh",
    "Greece": "gr",
    "Grenada": "gd",
    "Guatemala": "gt",
    "Guinea": "gn",
    "Guinea-Bissau": "gw",
    "Guyana": "gy",
    "Haiti": "ht",
    "Honduras": "hn",
    "Hungary": "hu",
    "Iceland": "is",
    "India": "in",
    "Indonesia": "id",
    "Iran": "ir",
    "Iraq": "iq",
    "Ireland": "ie",
    "Israel": "il",
    "Italy": "it",
    "Jamaica": "jm",
    "Japan": "jp",
    "Jordan": "jo",
    "Kazakhstan": "kz",
    "Kenya": "ke",
    "Kiribati": "ki",
    "Kuwait": "kw",
    "Kyrgyzstan": "kg",
    "Laos": "la",
    "Latvia": "lv",
    "Lebanon": "lb",
    "Lesotho": "ls",
    "Liberia": "lr",
    "Libya": "ly",
    "Liechtenstein": "li",
    "Lithuania": "lt",
    "Luxembourg": "lu",
    "Madagascar": "mg",
    "Malawi": "mw",
    "Malaysia": "my",
    "Maldives": "mv",
    "Mali": "ml",
    "Malta": "mt",
    "Marshall Islands": "mh",
    "Mauritania": "mr",
    "Mauritius": "mu",
    "Mexico": "mx",
    "Micronesia": "fm",
    "Moldova": "md",
    "Monaco": "mc",
    "Mongolia": "mn",
    "Montenegro": "me",
    "Morocco": "ma",
    "Mozambique": "mz",
    "Myanmar": "mm",
    "Namibia": "na",
    "Nauru": "nr",
    "Nepal": "np",
    "Netherlands": "nl",
    "New Zealand": "nz",
    "Nicaragua": "ni",
    "Niger": "ne",
    "Nigeria": "ng",
    "North Korea": "kp",
    "North Macedonia": "mk",
    "Norway": "no",
    "Oman": "om",
    "Pakistan": "pk",
    "Palau": "pw",
    "Panama": "pa",
    "Papua New Guinea": "pg",
    "Paraguay": "py",
    "Peru": "pe",
    "Philippines": "ph",
    "Poland": "pl",
    "Portugal": "pt",
    "Qatar": "qa",
    "Romania": "ro",
    "Russia": "ru",
    "Rwanda": "rw",
    "Saint Kitts and Nevis": "kn",
    "Saint Lucia": "lc",
    "Saint Vincent and the Grenadines": "vc",
    "Samoa": "ws",
    "San Marino": "sm",
    "Sao Tome and Principe": "st",
    "Saudi Arabia": "sa",
    "Senegal": "sn",
    "Serbia": "rs",
    "Seychelles": "sc",
    "Sierra Leone": "sl",
    "Singapore": "sg",
    "Slovakia": "sk",
    "Slovenia": "si",
    "Solomon Islands": "sb",
    "Somalia": "so",
    "South Africa": "za",
    "South Korea": "kr",
    "Spain": "es",
    "Sri Lanka": "lk",
    "Sudan": "sd",
    "Suriname": "sr",
    "Sweden": "se",
    "Switzerland": "ch",
    "Syria": "sy",
    "Taiwan": "tw",
    "Tajikistan": "tj",
    "Tanzania": "tz",
    "Thailand": "th",
    "Togo": "tg",
    "Tonga": "to",
    "Trinidad and Tobago": "tt",
    "Tunisia": "tn",
    "Turkey": "tr",
    "Turkmenistan": "tm",
    "Tuvalu": "tv",
    "Uganda": "ug",
    "Ukraine": "ua",
    "United Arab Emirates": "ae",
    "United Kingdom": "gb",
    "United States": "us",
    "Uruguay": "uy",
    "Uzbekistan": "uz",
    "Vanuatu": "vu",
    "Vatican City": "va",
    "Venezuela": "ve",
    "Vietnam": "vn",
    "Yemen": "ye",
    "Zambia": "zm",
    "Zimbabwe": "zw" 
}

@app.post("/check-rank")
async def check_rank(domain: str = Form(...), keywords: str = Form(...), country: str = Form(...)):
    # Limit the number of keywords to 500
    keyword_list = keywords.split(",")[:500]
    country_code = country_codes.get(country, "us")  # Default to US if country is not found
    result_data = []

    # Extract domain name from the given URL
    parsed_url = urlparse(domain)
    domain_name = parsed_url.netloc if parsed_url.netloc else domain

    # Create a new workbook for Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = domain_name  # Set sheet title as domain name

    # Add domain and country name to the first two rows
    current_date = datetime.now().strftime("%Y-%m-%d")
    sheet.cell(row=1, column=1, value="Domain").font = Font(bold=True)
    sheet.cell(row=1, column=2, value=domain_name)
    sheet.cell(row=2, column=1, value="Country Code").font = Font(bold=True)
    sheet.cell(row=2, column=2, value=country)
    sheet.cell(row=3, column=1, value="Date").font = Font(bold=True)
    sheet.cell(row=3, column=2, value=current_date)

    # Set up headers for the Excel sheet
    headers = ["Keyword", "Rank", "Page Number"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Fetch rankings for each keyword (up to 500 keywords)
    for row_num, keyword in enumerate(keyword_list, start=5):
        keyword = keyword.strip()
        search_url = f"https://serpapi.com/search?api_key={SERP_API_KEY}&q={keyword}&hl=en&gl={country_code}&num=100"
        try:
            response = requests.get(search_url)
            response.raise_for_status()
            data = response.json()

            # Extract ranking and page number information
            rank = None
            page_number = None
            for i, result in enumerate(data.get("organic_results", []), start=1):
                if domain in result.get("link", ""):
                    rank = i
                    page_number = (i - 1) // 10 + 1  # Page number calculation (10 results per page)
                    break

            # Write data to the Excel sheet
            sheet.cell(row=row_num, column=1, value=keyword)
            sheet.cell(row=row_num, column=2, value=rank or "Not Found")
            sheet.cell(row=row_num, column=3, value=page_number or "Not Found")

            result_data.append({
                "keyword": keyword,
                "rank": rank or "Not Found",
                "page_number": page_number or "Not Found"
            })
        except Exception as e:
            sheet.cell(row=row_num, column=1, value=keyword)
            sheet.cell(row=row_num, column=2, value=f"Error: {str(e)}")
            sheet.cell(row=row_num, column=3, value="Error")
            result_data.append({
                "keyword": keyword,
                "rank": f"Error: {str(e)}",
                "page_number": "Error"
            })

    # Save the Excel file with the domain name as its filename
    file_name = f"{domain_name.replace('.', '_')}_rankings.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)
    workbook.save(file_path)

    return {"results": result_data, "file_url": f"/files/{file_name}"}

@app.get("/download-ranking/{file_name}")
async def download_ranking_file(file_name: str):
    file_path = os.path.join(OUTPUT_DIR, file_name)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Ranking file not found.")
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=file_name)
# Signup Endpoint
@app.post("/signup/")
def signup(user: UserCreate = Body(...), db: Session = Depends(get_db)):
    existing_user = db.query(User).filter(
        (User.email_id == user.email_id) |
        (User.phone_number == user.phone_number) |
        (User.login_id == user.email_id.split('@')[0])
    ).first()

    if existing_user:
        if existing_user.email_id == user.email_id:
            raise HTTPException(status_code=400, detail="Email ID already registered")
        if existing_user.phone_number == user.phone_number:
            raise HTTPException(status_code=400, detail="Phone number already registered")
        if existing_user.login_id == user.email_id.split('@')[0]:
            raise HTTPException(status_code=400, detail="Login ID already taken")
    hashed_password = pwd_context.hash(user.password)
    login_id = user.email_id.split('@')[0]
    db_user = User(
        name=user.name,
        email_id=user.email_id,
        phone_number=user.phone_number,
        password_hash=hashed_password,
        login_id=login_id
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return {"message": "User created successfully", "login_id": login_id}

# Login Endpoint
@app.post("/login/")
def login(user: LoginSchema, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.email_id == user.email_id).first()
    if not db_user or not pwd_context.verify(user.password, db_user.password_hash):
        raise HTTPException(status_code=400, detail="Invalid credentials")
    return {"message": "Login successful", "email_id": db_user.email_id}
# Project Management
@app.post("/projects/")
def create_project(project: ProjectCreate, db: Session = Depends(get_db)):
    print("Received data:", project.dict()) 
    user = db.query(User).filter(User.id == project.user_id).first()  # Corrected query

    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db_project = Project(
        project_name=project.project_name,
        project_description=project.project_description,
        user_id=user.id  # Correctly assign user_id
    )
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    
    return {"message": "Project created successfully", "project_id": db_project.id}

@app.get("/projects/")
def get_projects(username: str,db: Session = Depends(get_db)):
    user = db.query(User).filter(User.name == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    projects = db.query(Project).filter(Project.user_id == user.id).all()
    return projects
def delete_project(db: Session, project_id: int):
    try:
        # Delete URLs, Keywords, and Ranks associated with the project
        db.query(URL).filter(URL.project_id == project_id).delete(synchronize_session="fetch")
        db.query(Keyword).filter(Keyword.project_id == project_id).delete(synchronize_session="fetch")
        db.query(Rank).filter(Rank.project_id == project_id).delete(synchronize_session="fetch")

        # Now delete the project itself
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        db.delete(project)
        db.commit()
        
        return {"message": "Project deleted successfully"}

    except Exception as e:
        db.rollback()  # Rollback in case of failure
        raise HTTPException(status_code=500, detail=f"Error deleting project: {str(e)}")

# URL & Keyword Management
scheduler = BackgroundScheduler()

@app.put("/edit-project/{project_id}")
async def edit_project(
    project_id: int,
    project_name: str = Form(...),
    project_description: str = Form(...),
    db: Session = Depends(get_db)
):
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return {"error": "Project not found"}
        
        project.project_name = project_name
        project.project_description = project_description
        db.commit()
        db.refresh(project)
        
        return {"message": "Project updated successfully", "project_id": project.id}
    except Exception as e:
        return {"error": str(e)}

SERP_API_KEY = "c833c9f303031dfdb298ae50981f8465d3d085aa263e5948923eff4ab2e9244c"

def get_country_code(country_name: str) -> str:
    try:
        country_obj = pycountry.countries.lookup(country_name)
        return country_obj.alpha_2.lower()  # Convert to lowercase
    except LookupError:
        logging.warning(f"⚠️ Invalid country name: {country_name}, defaulting to 'in'")
        return "in"  # Default to India if invalid

def get_country_details(country_name: str):
    try:
        country_obj = pycountry.countries.lookup(country_name)
        return country_obj.name, country_obj.alpha_2.lower()  # Full name & code
    except LookupError:
        logging.warning(f"⚠️ Invalid country name: {country_name}, defaulting to 'India'")
        return "India", "in"  # Default to India

@app.post("/track-rank")
async def track_rank(
    project_id: int = Form(...),
    domain: str = Form(...),
    keywords: str = Form(...),
    country: str = Form(...),
    db: Session = Depends(get_db)
):
    try:
        keyword_list = [k.strip() for k in keywords.split(",")[:500]]
        country_name, country_code = get_country_details(country.strip())  # Get full name & code

        # Ensure the URL exists in the project
        existing_url = db.query(URL).filter(URL.url == domain, URL.project_id == project_id).first()
        if not existing_url:
            existing_url = URL(url=domain, project_id=project_id, country=country_code)
            db.add(existing_url)
            db.commit()
            db.refresh(existing_url)

        results = []
        for keyword in keyword_list:
            # Check if keyword exists
            existing_keyword = db.query(Keyword).filter(
                Keyword.keyword == keyword,
                Keyword.url_id == existing_url.id
            ).first()

            if not existing_keyword:
                existing_keyword = Keyword(
                    keyword=keyword,
                    project_id=project_id,
                    url_id=existing_url.id
                )
                db.add(existing_keyword)
                db.commit()
                db.refresh(existing_keyword)

            # Fetch latest rank if available
            latest_rank = db.query(Rank).filter(
                Rank.keyword_id == existing_keyword.id,
                Rank.url_id == existing_url.id,
                Rank.project_id == project_id
            ).order_by(Rank.date.desc()).first()

            if latest_rank:
                results.append({
                    "keyword": keyword,
                    "rank": latest_rank.ranks,
                    "page_number": latest_rank.page_number,
                    "country": country_name,  
                    "date": latest_rank.date.strftime("%Y-%m-%d")
                })
                continue

            # Fetch Rank from SerpAPI
            search_url = f"https://serpapi.com/search?api_key={SERP_API_KEY}&q={quote_plus(keyword)}&hl=en&gl={country_code}&num=100"
            response = requests.get(search_url)

            if response.status_code == 200:
                data = response.json()
                rank = -1
                page_number = -1

                if "organic_results" in data:
                    for i, result in enumerate(data["organic_results"], start=1):
                        parsed_result_url = urlparse(result.get("link", "")).netloc.replace("www.", "")
                        parsed_stored_url = urlparse(domain).netloc.replace("www.", "")

                        if parsed_stored_url in parsed_result_url:
                            rank = i
                            page_number = (i - 1) // 10 + 1
                            break

                # Store rank in database
                new_rank = Rank(
                    url_id=existing_url.id,
                    keyword_id=existing_keyword.id,
                    project_id=project_id,
                    ranks=rank,
                    page_number=page_number,
                    country=country_code,
                    date=datetime.utcnow()
                )
                db.add(new_rank)
                db.commit()

                results.append({
                    "keyword": keyword,
                    "rank": rank,
                    "page_number": page_number,
                    "country": country_name,  
                    "date": datetime.utcnow().strftime("%Y-%m-%d")
                })
            else:
                results.append({
                    "keyword": keyword,
                    "error": f"Failed to fetch ranking: {response.status_code}"
                })

        return {"data": results}
    except Exception as e:
        logging.error(f"Error in /track-rank: {str(e)}")
        return {"error": str(e)}

#  Scheduler to run daily at 6:00 AM
def update_ranks():
    db = SessionLocal()
    urls = db.query(URL).all()

    for url in urls:
        for keyword in url.keywords:
            print(f" Checking Rank for: {url.url} - {keyword.keyword}")

            country_name, country_code = get_country_details(url.country if url.country else "United States")

            search_url = f"https://serpapi.com/search?api_key={SERP_API_KEY}&q={keyword.keyword}&hl=en&gl={country_code}&num=100"
            response = requests.get(search_url).json()

            rank = None
            page_number = None
            for i, result in enumerate(response.get("organic_results", []), start=1):
                if url.url in result.get("link", ""):
                    rank = i
                    page_number = (i - 1) // 10 + 1
                    print(f" Rank Found: {rank} (Page {page_number})")
                    break

            if rank is None:
                print(f" No Rank Found for {keyword.keyword}")

            # Store rank in DB
            new_rank = Rank(
                url_id=url.id,
                keyword_id=keyword.id,
                project_id=url.project_id,
                ranks=rank or -1,
                page_number=page_number or -1,
                country=country_code,
                date=datetime.utcnow()
            )
            db.add(new_rank)
            db.commit()
            print(f"Rank inserted: {url.url} - {keyword.keyword} => Rank: {rank} (Page {page_number})")

    db.close()

scheduler.add_job(update_ranks, 'cron', hour=6, minute=0)
scheduler.start()
print("Scheduler started! Runs daily at 6:00 AM.")

@app.post("/manual-update-ranks")
def manual_update_ranks():
    try:
        update_ranks()
        return {"message": "Ranks updated manually"}
    except Exception as e:
        return {"error": str(e)}

@app.get("/get-ranks")
async def get_ranks(project_id: int, db: Session = Depends(get_db)):
    try:
        ranks = db.query(Rank).filter(Rank.project_id == project_id).all()

        if not ranks:
            return {"message": "No rank data found"}

        result = []
        for rank in ranks:
            url = db.query(URL).filter(URL.id == rank.url_id).first()
            keyword = db.query(Keyword).filter(Keyword.id == rank.keyword_id).first()

            # Convert stored country code back to full country name
            country_name, _ = get_country_details(rank.country)

            result.append({
                "url": url.url if url else "Deleted URL",
                "keyword": keyword.keyword if keyword else "Deleted Keyword",
                "ranks": rank.ranks,
                "page_number": rank.page_number,
                "country": country_name,  
                "date": rank.date.strftime("%Y-%m-%d")
            })

        return result
    except Exception as e:
        return {"error": str(e)}

@app.post("/add-keywords")
async def add_keywords(
    project_id: int = Form(...),
    domain: str = Form(...),
    keywords: str = Form(...),
    country: str = Form(...),
    db: Session = Depends(get_db)
):
    try:
        keyword_list = [k.strip() for k in keywords.split(",")[:500]]
        country_lower = country.lower().strip()
        google_domain = country_domains.get(country_lower, "google.com")

        # Ensure the URL exists
        existing_url = db.query(URL).filter(URL.url == domain, URL.project_id == project_id).first()
        if not existing_url:
            return {"error": "URL not found in the project"}

        serp_api_key = "serp_api_key"
        added_keywords = []

        for keyword in keyword_list:
            # Check if keyword already exists for this project & URL
            existing_keyword = db.query(Keyword).filter(
                Keyword.keyword == keyword,
                Keyword.project_id == project_id,
                Keyword.url_id == existing_url.id  
            ).first()

            if not existing_keyword:
                new_keyword = Keyword(
                    keyword=keyword,
                    project_id=project_id,
                    url_id=existing_url.id  
                )
                db.add(new_keyword)
                db.commit()
                db.refresh(new_keyword)
                added_keywords.append(new_keyword)

                # Fetch rank immediately using SerpAPI
                search_url = f"https://serpapi.com/search?api_key={serp_api_key}&q={quote_plus(keyword)}&hl=en&gl={country_lower}&num=100&google_domain={google_domain}"
                response = requests.get(search_url)

                if response.status_code == 200:
                    data = response.json()
                    rank = -1
                    page_number = -1

                    if "organic_results" in data:
                        for i, result in enumerate(data["organic_results"], start=1):
                            result_url = result.get("link", "").strip()
                            if not result_url:
                                continue

                            parsed_result_url = urlparse(result_url).netloc.replace("www.", "")
                            parsed_stored_url = urlparse(domain).netloc.replace("www.", "")

                            if parsed_stored_url in parsed_result_url:
                                rank = i
                                page_number = (i - 1) // 10 + 1
                                break

                    new_rank = Rank(
                        url_id=existing_url.id,
                        keyword_id=new_keyword.id,
                        project_id=project_id,
                        ranks=rank,
                        page_number=page_number,
                        country=country_lower,
                        date=datetime.utcnow()
                    )
                    db.add(new_rank)
                    db.commit()

        return {"message": f"{len(added_keywords)} new keywords added"}

    except Exception as e:
        return {"error": str(e)}

# Run Migrations
Base.metadata.create_all(bind=engine)


@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/get-keywords")
async def get_keywords(
    project_id: int,
    domain: str,
    db: Session = Depends(get_db)
):
    try:
        # Ensure the URL exists
        existing_url = db.query(URL).filter(URL.url == domain, URL.project_id == project_id).first()
        if not existing_url:
            return {"error": "URL not found in the project"}

        # Fetch keywords for the given project and URL
        keywords = db.query(Keyword).filter(
            Keyword.project_id == project_id,
            Keyword.url_id == existing_url.id
        ).all()

        if not keywords:
            return {"message": "No keywords found for this project and URL"}

        return [
            {"keyword_id": k.id, "keyword": k.keyword, "url_id": k.url_id} for k in keywords
        ]

    except Exception as e:
        return {"error": str(e)}

@app.delete("/delete-keyword/{keyword_id}")
async def delete_keyword(
    keyword_id: int,
    db: Session = Depends(get_db)
):
    try:
        # Find the keyword
        keyword = db.query(Keyword).filter(Keyword.id == keyword_id).first()
        if not keyword:
            return {"error": "Keyword not found"}

        # Delete related rankings
        db.query(Rank).filter(Rank.keyword_id == keyword_id).delete()

        # Delete the keyword
        db.delete(keyword)
        db.commit()

        return {"message": "Keyword and associated rankings deleted successfully"}

    except Exception as e:
        db.rollback()
        return {"error": str(e)}

@app.get("/live-rank")
async def get_live_rank(project_id: int, db: Session = Depends(get_db)):
    try:
        # Step 1: Get keywords and URLs for the project
        keywords = db.query(Keyword).filter(Keyword.project_id == project_id).all()
        urls = db.query(URL).filter(URL.project_id == project_id).all()
        project = db.query(Project).filter(Project.id == project_id).first()

        if not project or not keywords or not urls:
            raise HTTPException(status_code=404, detail="Project, URLs, or Keywords not found")

        country = project.country  # assuming you store it (like 'in' or 'us')

        results = []
        for keyword in keywords:
            for url in urls:
                # Step 2: Fetch live rank from SerpAPI
                serp_api_key = "serp_api_key"  
                params = {
                    "api_key": serp_api_key,
                    "q": keyword.keyword,
                    "gl": country,
                    "hl": "en",
                    "num": "100",
                    "engine": "google"
                }

                response = requests.get("https://serpapi.com/search", params=params)
                data = response.json()

                # Step 3: Search through results to find rank
                position = None
                page_number = None
                for index, result in enumerate(data.get("organic_results", []), start=1):
                    if url.url in result.get("link", ""):
                        position = index
                        page_number = (index - 1) // 10 + 1
                        break

                results.append({
                    "keyword": keyword.keyword,
                    "url": url.url,
                    "rank": position if position else "Not found",
                    "page_number": page_number if page_number else "N/A",
                    "checked_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                })

        return {"project_id": project_id, "live_results": results}

    except Exception as e:
        return {"error": str(e)}

@app.get("/get-1-day-rank")
async def get_1_day_rank(
    project_id: int,
    db: Session = Depends(get_db)
):
    try:
        # Set date range: start of yesterday to start of today
        today = datetime.utcnow().date()
        start_of_yesterday = datetime.combine(today - timedelta(days=1), datetime.min.time())
        start_of_today = datetime.combine(today, datetime.min.time())

        rank_data = (
            db.query(Rank, Keyword, URL)
            .join(Keyword, Rank.keyword_id == Keyword.id)
            .join(URL, Rank.url_id == URL.id)
            .filter(
                Rank.project_id == project_id,
                Rank.date >= start_of_yesterday,
                Rank.date < start_of_today  # Exclude today's data
            )
            .order_by(Rank.date.desc())
            .all()
        )

        results = []
        for rank, keyword, url in rank_data:
            results.append({
                "keyword": keyword.keyword,
                "domain": url.url,
                "rank": rank.ranks,
                "page_number": rank.page_number,
                "country": rank.country,
                "date": rank.date.strftime("%Y-%m-%d")
            })

        return {"data": results}

    except Exception as e:
        logging.error(f"Error in /get-1-day-rank: {str(e)}")
        return {"error": str(e)}

@app.get("/get-7-days-rank")
async def get_7_days_rank(
    project_id: int,
    db: Session = Depends(get_db)
):
    try:
        seven_days_ago = datetime.utcnow() - timedelta(days=7)

 
        rank_data = (
            db.query(Rank, Keyword, URL)
            .join(Keyword, Rank.keyword_id == Keyword.id)  # Explicit join
            .join(URL, Rank.url_id == URL.id)  # Explicit join
            .filter(
                Rank.project_id == project_id,
                Rank.date >= seven_days_ago
            )
            .order_by(Rank.date.desc())
            .all()
        )

        results = []
        for rank, keyword, url in rank_data:
            results.append({
                "keyword": keyword.keyword,
                "domain": url.url,  # Correctly getting URL
                "rank": rank.ranks,
                "page_number": rank.page_number,
                "country": rank.country,
                "date": rank.date.strftime("%Y-%m-%d")
            })

        return {"data": results}

    except Exception as e:
        logging.error(f"Error in /get-7-days-rank: {str(e)}")
        return {"error": str(e)}

@app.get("/get-30-days-rank")
async def get_30_days_rank(
    project_id: int,
    db: Session = Depends(get_db)
):
    try:
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)

        # Step 1: Fetch rank data from the last 30 days with explicit joins
        rank_data = (
            db.query(Rank, Keyword, URL)
            .select_from(Rank)
            .join(Keyword, Rank.keyword_id == Keyword.id)
            .join(URL, Rank.url_id == URL.id)
            .filter(
                Rank.project_id == project_id,
                Rank.date >= thirty_days_ago
            )
            .order_by(Rank.date.desc())
            .all()
        )

        if rank_data:
            results = []
            for rank, keyword, url in rank_data:
                results.append({
                    "keyword": keyword.keyword,
                    "domain": url.url,
                    "rank": rank.ranks,
                    "page_number": rank.page_number,
                    "country": rank.country,
                    "date": rank.date.strftime("%Y-%m-%d")
                })
            return {"data": results}

        # Step 2: Fallback to latest tracked data
        last_tracked_data = (
            db.query(Rank, Keyword, URL)
            .select_from(Rank)
            .join(Keyword, Rank.keyword_id == Keyword.id)
            .join(URL, Rank.url_id == URL.id)
            .filter(Rank.project_id == project_id)
            .order_by(Rank.date.desc())
            .all()
        )

        if last_tracked_data:
            results = []
            for rank, keyword, url in last_tracked_data:
                results.append({
                    "keyword": keyword.keyword,
                    "domain": url.url,
                    "rank": rank.ranks,
                    "page_number": rank.page_number,
                    "country": rank.country,
                    "date": rank.date.strftime("%Y-%m-%d")
                })
            return {
                "data": results,
                "message": "No rank in last 30 days. Returning the last tracked data."
            }

        # Step 3: No data at all
        return {"data": [], "message": "No ranking data available for this project."}

    except Exception as e:
        logging.error(f"Error in /get-30-days-rank: {str(e)}")
        return {"error": str(e)}
from typing import List
from fastapi import Body

@app.delete("/delete-keywords")
async def delete_keywords(
    keyword_ids: List[int] = Body(...),  # Accept list of IDs in the request body
    db: Session = Depends(get_db)
):
    try:
        # Get the keywords from DB
        keywords = db.query(Keyword).filter(Keyword.id.in_(keyword_ids)).all()
        
        if not keywords:
            raise HTTPException(status_code=404, detail="No keywords found with the given IDs")

        # Delete related rankings
        db.query(Rank).filter(Rank.keyword_id.in_(keyword_ids)).delete(synchronize_session=False)

        # Delete the keywords
        for keyword in keywords:
            db.delete(keyword)

        db.commit()

        return {"message": "Keywords and associated rankings deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
class URLDeleteRequest(BaseModel):
    url_ids: list[int]

@app.delete("/delete-urls")
async def delete_urls(request: URLDeleteRequest, db: Session = Depends(get_db)):
    try:
        # Delete all related ranks for these URLs
        db.query(Rank).filter(Rank.url_id.in_(request.url_ids)).delete(synchronize_session=False)

        # Delete the URLs themselves
        db.query(URL).filter(URL.id.in_(request.url_ids)).delete(synchronize_session=False)

        db.commit()
        return {"message": "URLs and associated rankings deleted successfully"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/live-rank")
async def get_live_rank(project_id: int, db: Session = Depends(get_db)):
    try:
        # Fetch project, keywords, and URLs
        project = db.query(Project).filter(Project.id == project_id).first()
        keywords = db.query(Keyword).filter(Keyword.project_id == project_id).all()
        urls = db.query(URL).filter(URL.project_id == project_id).all()

        if not project or not keywords or not urls:
            raise HTTPException(status_code=404, detail="Project, URLs, or Keywords not found")

        if not project.country:
            raise HTTPException(status_code=400, detail="Project does not have a country assigned")

        country = project.country.lower()

        results = []
        for keyword in keywords:
            for url in urls:
                serp_api_key = "c833c9f303031dfdb298ae50981f8465d3d085aa263e5948923eff4ab2e9244c"
                params = {
                    "api_key": serp_api_key,
                    "q": keyword.keyword,
                    "gl": country,
                    "hl": "en",
                    "num": "100",
                    "engine": "google"
                }

                response = requests.get("https://serpapi.com/search", params=params)
                data = response.json()

                position = None
                page_number = None
                for index, result in enumerate(data.get("organic_results", []), start=1):
                    if url.url in result.get("link", ""):
                        position = index
                        page_number = (index - 1) // 10 + 1
                        break

                results.append({
                    "keyword": keyword.keyword,
                    "url": url.url,
                    "rank": position if position else "Not found",
                    "page_number": page_number if page_number else "N/A",
                    "checked_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                })

        return {"project_id": project_id, "country": country, "live_results": results}

    except Exception as e:
        return {"error": str(e)}

Required packages 
Pip install fastapi, uvicorn[standard] ,sqlalchemy ,passlib[bcrypt] ,pydantic ,apscheduler ,bcrypt ,requests
Xlsxwriter ,openpyxl ,jinja2 ,python-multipart ,iso3166 ,pycountry



